"""发票管理 API 路由"""

import copy
import logging
from datetime import date, datetime
import io
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from sqlalchemy import or_
from sqlalchemy.orm import Session, selectinload

from ..database import get_db
from ..models.invoice import Category, Counterpart, Invoice, InvoiceDetail, InvoiceFile
from ..resource_path import get_template_path
from ..schemas.invoice import (
    Category as CategorySchema,
)
from ..schemas.invoice import (
    CategoryCreate,
    CounterpartCreate,
    InvoiceCreate,
    InvoiceUpdate,
    PaginatedResponse,
)
from ..schemas.invoice import (
    Counterpart as CounterpartSchema,
)
from ..schemas.invoice import (
    InvoiceResponse as InvoiceSchema,
)
from ..schemas.invoice import (
    InvoiceFileResponse as InvoiceFileSchema,
)
from ..services.file_storage import (
    MAGIC_LEN,
    _detect_mime,
    _MIME_TO_EXTS,
    delete_file,
    retrieve_file,
    store_file,
    upload_invoice_file,
)

router = APIRouter(prefix="/api", tags=["发票管理"])
logger = logging.getLogger(__name__)


# ==================== 发票 CRUD ====================

@router.post("/invoices", response_model=InvoiceSchema, status_code=201)
def create_invoice(invoice_data: InvoiceCreate, db: Session = Depends(get_db)):
    """创建发票（含明细）"""
    try:
        invoice = Invoice(**invoice_data.model_dump(exclude={"details"}))
        db.add(invoice)

        for detail_data in invoice_data.details:
            detail = InvoiceDetail(invoice=invoice, **detail_data.model_dump())
            db.add(detail)

        db.commit()
        db.refresh(invoice)
        return invoice
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="创建发票失败")


@router.get("/invoices", response_model=PaginatedResponse[InvoiceSchema])
def list_invoices(
    invoice_number: str | None = Query(None),
    invoice_code: str | None = Query(None),
    invoice_type: str | None = Query(None),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    min_amount: float | None = Query(None),
    max_amount: float | None = Query(None),
    counterpart_id: int | None = Query(None),
    category_id: int | None = Query(None),
    search_text: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """查询发票列表，支持多条件筛选"""
    query = db.query(Invoice).options(
        selectinload(Invoice.counterpart),
        selectinload(Invoice.category),
        selectinload(Invoice.details),
    )

    if invoice_number:
        query = query.filter(Invoice.invoice_number.like(f"%{invoice_number}%"))
    if invoice_code:
        query = query.filter(Invoice.invoice_code.like(f"%{invoice_code}%"))
    if invoice_type:
        query = query.filter(Invoice.invoice_type == invoice_type)
    if start_date:
        query = query.filter(Invoice.invoice_date >= start_date)
    if end_date:
        query = query.filter(Invoice.invoice_date <= end_date)
    if min_amount is not None:
        query = query.filter(Invoice.total_with_tax >= min_amount)
    if max_amount is not None:
        query = query.filter(Invoice.total_with_tax <= max_amount)
    if counterpart_id:
        query = query.filter(Invoice.counterpart_id == counterpart_id)
    if category_id:
        query = query.filter(Invoice.category_id == category_id)
    if search_text:
        escaped = search_text.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
        # JOIN counterpart 以支持按单位名称搜索
        query = query.outerjoin(Counterpart, Invoice.counterpart_id == Counterpart.id)
        query = query.filter(
            or_(
                Invoice.invoice_number.like(f"%{escaped}%", escape="\\"),
                Invoice.invoice_code.like(f"%{escaped}%", escape="\\"),
                Invoice.remark.like(f"%{escaped}%", escape="\\"),
                Invoice.raw_text.like(f"%{escaped}%", escape="\\"),
                Invoice.check_code.like(f"%{escaped}%", escape="\\"),
                Counterpart.name.like(f"%{escaped}%", escape="\\"),
            )
        )

    total = query.count()
    items = query.order_by(Invoice.invoice_date.desc()).offset(skip).limit(limit).all()

    return {"items": items, "total": total}


@router.get("/invoices/{invoice_id}", response_model=InvoiceSchema)
def get_invoice(invoice_id: int, db: Session = Depends(get_db)):
    """获取单张发票详情"""
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    return invoice


@router.put("/invoices/{invoice_id}", response_model=InvoiceSchema)
def update_invoice(invoice_id: int, invoice_data: InvoiceUpdate, db: Session = Depends(get_db)):
    """更新发票信息"""
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    try:
        # 更新发票主字段
        update_dict = invoice_data.model_dump(exclude={"details"}, exclude_unset=True)
        for key, value in update_dict.items():
            setattr(invoice, key, value)

        # 更新明细（如果提供了）
        if invoice_data.details is not None:
            # 删除旧明细
            db.query(InvoiceDetail).filter(InvoiceDetail.invoice_id == invoice_id).delete()
            # 添加新明细
            for detail_data in invoice_data.details:
                detail = InvoiceDetail(invoice_id=invoice_id, **detail_data.model_dump())
                db.add(detail)

        db.commit()
        db.refresh(invoice)
        return invoice
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="更新发票失败")


@router.delete("/invoices/{invoice_id}", status_code=204)
def delete_invoice(invoice_id: int, db: Session = Depends(get_db)):
    """删除发票（含关联明细和文件）"""
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # 先收集待删文件记录（不立即删物理文件，避免 DB 回滚后文件已删）
    file_records = list(invoice.files)

    db.delete(invoice)
    db.commit()

    # 事务提交后再清理物理文件；任一失败不影响已提交的删除结果
    for file in file_records:
        try:
            delete_file(file)
        except Exception:
            # 文件删除失败不影响主流程，留待后续清理任务
            pass


# ==================== 发票文件管理 ====================

@router.post("/invoices/{invoice_id}/files", response_model=InvoiceFileSchema, status_code=201)
def upload_invoice_file(invoice_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """上传发票原文件"""
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    # MIME 嗅探：先读前 16 字节检测魔数，再 seek(0) 复位供流式拷贝使用
    head = file.file.read(MAGIC_LEN)
    try:
        file.file.seek(0)
    except OSError:
        # 某些流式源不可 seek，回退到一次性读取（极少触发）
        file.file.seek(0, 2)
        remaining = file.file.tell()
        file.file.seek(0)
        head = head[: min(len(head), remaining)]

    detected_mime = _detect_mime(head)
    if not detected_mime:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件类型或文件已损坏: {Path(file.filename).suffix.lower()}",
        )

    # 后缀必须与嗅探到的 MIME 匹配，防止「改后缀绕过」
    ext = Path(file.filename).suffix.lower().lstrip(".")
    allowed_exts = _MIME_TO_EXTS.get(detected_mime, set())
    if ext not in allowed_exts:
        raise HTTPException(
            status_code=400,
            detail=f"文件扩展名 {ext} 与实际类型 {detected_mime} 不匹配",
        )

    # 流式分块写入（CHUNK_SIZE=1MB），写入过程强制 MAX_FILE_SIZE 上限
    from ..config import settings
    max_bytes = settings.max_file_size_bytes
    try:
        storage_mode, file_size, file_path, blob_data = upload_invoice_file(
            file.file, file.filename, invoice_id, max_bytes
        )
    except Exception:
        raise HTTPException(status_code=500, detail="文件存储失败")

    # 创建文件记录
    invoice_file = InvoiceFile(
        invoice_id=invoice_id,
        file_name=file.filename,
        file_type=ext.upper(),
        file_size=file_size,
        storage_mode=storage_mode,
        file_path=file_path,
        blob_data=blob_data,
    )
    try:
        db.add(invoice_file)
        db.commit()
        db.refresh(invoice_file)
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="文件记录保存失败")

    return invoice_file


@router.get("/invoices/{invoice_id}/files", response_model=list[InvoiceFileSchema])
def list_invoice_files(invoice_id: int, db: Session = Depends(get_db)):
    """获取发票的所有原文件"""
    files = db.query(InvoiceFile).filter(InvoiceFile.invoice_id == invoice_id).all()
    return files


@router.get("/invoices/{invoice_id}/files/{file_id}/download")
def download_invoice_file(invoice_id: int, file_id: int, db: Session = Depends(get_db)):
    """下载发票原文件"""
    file_record = db.query(InvoiceFile).filter(
        InvoiceFile.id == file_id,
        InvoiceFile.invoice_id == invoice_id,
    ).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在")

    try:
        content = retrieve_file(
            file_record.storage_mode,
            file_record.file_path,
            file_record.blob_data,
        )
    except Exception:
        logger.exception("读取文件失败: file_id=%s", file_record.id)
        raise HTTPException(status_code=500, detail="文件读取失败")

    media_types = {
        "PDF": "application/pdf",
        "PNG": "image/png",
        "JPG": "image/jpeg",
        "JPEG": "image/jpeg",
    }
    media_type = media_types.get(file_record.file_type, "application/octet-stream")

    return FileResponse(
        io.BytesIO(content),
        media_type=media_type,
        filename=file_record.file_name,
    )


@router.delete("/invoices/{invoice_id}/files/{file_id}", status_code=204)
def delete_invoice_file(invoice_id: int, file_id: int, db: Session = Depends(get_db)):
    """删除发票原文件"""
    file_record = db.query(InvoiceFile).filter(
        InvoiceFile.id == file_id,
        InvoiceFile.invoice_id == invoice_id,
    ).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="文件不存在")

    delete_file(file_record)
    db.delete(file_record)
    db.commit()


# ==================== 分类管理 ====================

@router.post("/categories", response_model=CategorySchema, status_code=201)
def create_category(category_data: CategoryCreate, db: Session = Depends(get_db)):
    """创建消费分类"""
    category = Category(**category_data.model_dump())
    db.add(category)
    db.commit()
    db.refresh(category)
    return category


@router.get("/categories", response_model=list[CategorySchema])
def list_categories(db: Session = Depends(get_db)):
    """获取所有消费分类"""
    return db.query(Category).all()


@router.put("/categories/{category_id}", response_model=CategorySchema)
def update_category(category_id: int, category_data: CategoryCreate, db: Session = Depends(get_db)):
    """更新消费分类"""
    category = db.query(Category).filter(Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类不存在")
    for key, value in category_data.model_dump().items():
        setattr(category, key, value)
    db.commit()
    db.refresh(category)
    return category


@router.delete("/categories/{category_id}", status_code=204)
def delete_category(category_id: int, db: Session = Depends(get_db)):
    """删除消费分类"""
    category = db.query(Category).filter(Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类不存在")
    db.delete(category)
    db.commit()


# ==================== 对方单位管理 ====================

@router.post("/counterparts", response_model=CounterpartSchema, status_code=201)
def create_counterpart(counterpart_data: CounterpartCreate, db: Session = Depends(get_db)):
    """创建对方单位"""
    counterpart = Counterpart(**counterpart_data.model_dump())
    db.add(counterpart)
    db.commit()
    db.refresh(counterpart)
    return counterpart


@router.get("/counterparts", response_model=list[CounterpartSchema])
def list_counterparts(db: Session = Depends(get_db)):
    """获取所有对方单位"""
    return db.query(Counterpart).all()


@router.put("/counterparts/{counterpart_id}", response_model=CounterpartSchema)
def update_counterpart(counterpart_id: int, counterpart_data: CounterpartCreate, db: Session = Depends(get_db)):
    """更新对方单位"""
    counterpart = db.query(Counterpart).filter(Counterpart.id == counterpart_id).first()
    if not counterpart:
        raise HTTPException(status_code=404, detail="单位不存在")
    for key, value in counterpart_data.model_dump().items():
        setattr(counterpart, key, value)
    db.commit()
    db.refresh(counterpart)
    return counterpart


@router.delete("/counterparts/{counterpart_id}", status_code=204)
def delete_counterpart(counterpart_id: int, db: Session = Depends(get_db)):
    """删除对方单位"""
    counterpart = db.query(Counterpart).filter(Counterpart.id == counterpart_id).first()
    if not counterpart:
        raise HTTPException(status_code=404, detail="单位不存在")
    db.delete(counterpart)
    db.commit()


# ==================== 发票导出 ====================

# 模板路径通过 resource_path 模块统一管理，兼容开发/打包环境

TEMPLATE_PATH = get_template_path()

@router.post("/invoices/export")
def export_invoices(invoice_ids: list[int] = Body(...), db: Session = Depends(get_db)):
    """导出选中发票到 Excel（按模板格式）"""
    if not invoice_ids:
        raise HTTPException(status_code=400, detail="请选择至少一张发票")

    invoices = db.query(Invoice).options(
        selectinload(Invoice.category)
    ).filter(Invoice.id.in_(invoice_ids)).order_by(
        Invoice.invoice_date.asc()
    ).all()
    if not invoices:
        raise HTTPException(status_code=404, detail="未找到选中的发票")

    # —— 按分类汇总：同分类合并为一条，金额求和 ——
    cat_order = []          # 保持分类首次出现顺序
    cat_groups = {}         # category_name -> {amount, count, remarks, dates}
    for inv in invoices:
        cat_name = inv.category.name if (inv.category and inv.category.name) else "其他"
        if cat_name not in cat_groups:
            cat_groups[cat_name] = {"amount": 0.0, "count": 0, "remarks": [], "dates": []}
            cat_order.append(cat_name)
        cat_groups[cat_name]["amount"] += float(inv.total_with_tax or 0)
        cat_groups[cat_name]["count"] += 1
        if inv.remark:
            cat_groups[cat_name]["remarks"].append(inv.remark.strip())
        if inv.invoice_date:
            cat_groups[cat_name]["dates"].append(inv.invoice_date)

    # 构建汇总行列表
    summary_rows = []
    for cat_name in cat_order:
        g = cat_groups[cat_name]
        # B列日期：取最早日期，格式化为中文月份
        date_str = ""
        if g["dates"]:
            d = g["dates"][0]
            if isinstance(d, str):
                try:
                    d = datetime.strptime(d[:10], "%Y-%m-%d").date()
                except ValueError:
                    date_str = str(d)
            if isinstance(d, date):
                date_str = f"{d.month}月"
            elif not date_str:
                date_str = str(d)
        # E列备注：合并多条
        remark_str = "；".join(g["remarks"]) if g["remarks"] else ""
        summary_rows.append({
            "category": cat_name,
            "date_str": date_str,
            "amount": round(g["amount"], 2),
            "count": g["count"],
            "remark": remark_str,
        })

    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail=f"模板文件不存在: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # 找到表头行 (含 "No." 或 "内容" 的那一行)
    header_row = 1
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and str(cell.value).strip() in ("No.", "内容", "金额"):
                header_row = cell.row
                break

    data_start_row = header_row + 1

    # 收集样式模板行（取第一行数据行作为样式参考，但边框始终清空）
    template_styles = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=data_start_row, column=col_idx)
        template_styles[col_idx] = {
            "font": copy.copy(cell.font),
            "fill": copy.copy(cell.fill),
            "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
        }

    # 先找到合计行
    summary_row = None
    for row in range(data_start_row, ws.max_row + 1):
        for cell in ws[row]:
            if cell.value and "合计" in str(cell.value):
                summary_row = row
                break
        if summary_row:
            break

    # 先解除数据区的合并单元格
    merged_to_unmerge = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row >= data_start_row:
            merged_to_unmerge.append(str(mc))
    for mc_str in merged_to_unmerge:
        ws.unmerge_cells(mc_str)

    # 清除所有旧数据（包括旧合计行）
    clear_end = max(summary_row, ws.max_row) if summary_row else ws.max_row
    for row in range(data_start_row, clear_end + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    # 填充汇总数据（每个分类一行）
    # 定义细线边框样式
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for i, sr in enumerate(summary_rows):
        row = data_start_row + i

        ws.cell(row=row, column=1, value=i + 1)          # A: 序号
        ws.cell(row=row, column=2, value=sr["date_str"])  # B: 日期（月份）
        ws.cell(row=row, column=3, value=sr["category"])  # C: 内容（分类）
        ws.cell(row=row, column=4, value=sr["amount"])    # D: 金额（求和）
        ws.cell(row=row, column=5, value=sr["count"])     # E: 发票数量
        ws.cell(row=row, column=6, value=sr["remark"])    # F: 备注
        ws.cell(row=row, column=7, value=sr["category"])  # G: 费用类别

        # 应用样式（细线边框 + 居中 + 备注自动换行）
        for col_idx, style in template_styles.items():
            cell = ws.cell(row=row, column=col_idx)
            if style.get("font"):
                cell.font = copy.copy(style["font"])
            if style.get("fill"):
                cell.fill = copy.copy(style["fill"])
            cell.border = thin_border
            # F列（备注）居中+自动换行，其他列居中
            if col_idx == 6:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # D列（金额）和 E列（发票数量）保留数字格式，其他列用模板格式
            if col_idx not in (4, 5):
                cell.number_format = style.get("number_format", "@")

    # 写入新合计行
    summary_new_row = data_start_row + len(summary_rows)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=summary_new_row, column=col).value = None
    merged_unmerge = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row == summary_new_row or mc.max_row == summary_new_row:
            merged_unmerge.append(str(mc))
    for mc_str in merged_unmerge:
        ws.unmerge_cells(mc_str)
    ws.cell(row=summary_new_row, column=3, value="合计")
    ws.cell(row=summary_new_row, column=4, value=f"=SUM(D{data_start_row}:D{summary_new_row - 1})")
    ws.cell(row=summary_new_row, column=5, value=f"=SUM(E{data_start_row}:E{summary_new_row - 1})")
    for col_idx, style in template_styles.items():
        cell = ws.cell(row=summary_new_row, column=col_idx)
        cell.font = copy.copy(style["font"])
        cell.fill = copy.copy(style["fill"])
        cell.border = thin_border
        cell.alignment = copy.copy(style["alignment"])

    # 删除合计行之后的多余空行（模板预留的空白数据行）
    if summary_row and summary_row > summary_new_row:
        ws.delete_rows(summary_new_row + 1, summary_row - summary_new_row)

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    encoded_filename = quote("费用报销汇总.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )
